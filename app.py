from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import process as rf_process
    from rapidfuzz import fuzz as rf_fuzz
except Exception:  # pragma: no cover
    rf_process = None
    rf_fuzz = None


# =========================
# Decisions (based on your reference workbook)
# =========================
# 1) UI: Streamlit (local web UI)
# 2) No open-source LLM/model required by default because the reference workbook already provides
#    NAICS <-> EF mapping and (often) Category/Product <-> NAICS mapping.


REF_ENV_VAR = "SCOPE3_REF_XLSX_PATH"
_REF_DEFAULT_LOCAL = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "product wise sales and procurement details.xlsx"))


def resolve_default_reference_path() -> Optional[str]:
    """Resolve a default reference workbook path for local runs or deployments.

    Priority:
    1) Environment variable SCOPE3_REF_XLSX_PATH (mounted file in container/server)
    2) ../product wise sales and procurement details.xlsx (local dev layout)
    """
    env_path = os.environ.get(REF_ENV_VAR)
    if env_path and os.path.exists(env_path):
        return env_path
    if os.path.exists(_REF_DEFAULT_LOCAL):
        return _REF_DEFAULT_LOCAL
    return None


@dataclass
class ReferenceData:
    ef_table: pd.DataFrame
    naics_lookup: pd.DataFrame
    catprod_to_naics: Optional[pd.DataFrame]
    default_fx_inr_per_usd: Optional[float]


def _safe_float(x) -> Optional[float]:
    try:
        if pd.isna(x):
            return None
        return float(x)
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def load_reference(ref_path: str) -> ReferenceData:
    # EF & Conversion: we need NAICS Code, Title, and EF with margins (col G)
    ef = pd.read_excel(ref_path, sheet_name="EF & Conversion", engine="openpyxl")

    # Normalize column names (but keep original EF table for export)
    # Expected columns from your file:
    # - 2017 NAICS Code
    # - 2017 NAICS Title
    # - Supply Chain Emission Factors with Margins
    naics_code_col = None
    naics_title_col = None
    ef_margin_col = None
    for c in ef.columns:
        cl = str(c).strip().lower()
        if naics_code_col is None and "naics" in cl and "code" in cl:
            naics_code_col = c
        if naics_title_col is None and "naics" in cl and ("title" in cl or "name" in cl):
            naics_title_col = c
        if ef_margin_col is None and "with" in cl and "margin" in cl and "emission" in cl:
            ef_margin_col = c

    if naics_code_col is None or naics_title_col is None or ef_margin_col is None:
        raise ValueError(
            "Could not locate required columns in 'EF & Conversion' sheet. "
            "Expected NAICS Code, NAICS Title/Name, and EF with Margins columns."
        )

    naics_lookup = ef[[naics_code_col, naics_title_col, ef_margin_col]].copy()
    naics_lookup.columns = ["naics_code", "naics_name", "ef_kgco2e_per_usd_margin"]

    # Ensure NAICS code is numeric-like for stable lookups
    naics_lookup["naics_code"] = pd.to_numeric(naics_lookup["naics_code"], errors="coerce")

    # Optional mapping table: Category + Product -> NAICS Code
    catprod_to_naics = None
    try:
        m = pd.read_excel(
            ref_path,
            sheet_name="Calculation Sheet",
            header=5,  # header row inferred from your workbook preview
            usecols="B:D",
            engine="openpyxl",
        )
        m = m.dropna(how="all")
        # Expect columns: Category, Product, NAICS Code
        cols = [str(c).strip().lower() for c in m.columns]
        if len(cols) >= 3 and ("category" in cols[0]) and ("product" in cols[1]) and ("naics" in cols[2]):
            m.columns = ["category", "product", "naics_code"]
            m["naics_code"] = pd.to_numeric(m["naics_code"], errors="coerce")
            m = m.dropna(subset=["category", "product", "naics_code"]).drop_duplicates()
            catprod_to_naics = m
    except Exception:
        catprod_to_naics = None

    # Default FX from EF & Conversion!M1 (your file stores the FY average there)
    default_fx = None
    try:
        wb = load_workbook(ref_path, read_only=True, data_only=True)
        ws = wb["EF & Conversion"]
        default_fx = _safe_float(ws["M1"].value)
    except Exception:
        default_fx = None

    return ReferenceData(
        ef_table=ef,
        naics_lookup=naics_lookup,
        catprod_to_naics=catprod_to_naics,
        default_fx_inr_per_usd=default_fx,
    )


def infer_spend_column(df: pd.DataFrame) -> str:
    numeric_cols = []
    for c in df.columns:
        s = pd.to_numeric(df[c], errors="coerce")
        if s.notna().sum() >= max(5, int(0.2 * len(df))):
            numeric_cols.append(c)

    if not numeric_cols:
        raise ValueError("No numeric spend-like column detected.")

    def score(col: str) -> float:
        name = str(col).lower()
        name_score = 0
        for kw in ["amount", "value", "spend", "cost", "total", "inward", "purchase", "invoice"]:
            if kw in name:
                name_score += 3
        # prefer columns with higher fill-rate and positive values
        s = pd.to_numeric(df[col], errors="coerce")
        fill = s.notna().mean()
        pos = (s.fillna(0) > 0).mean()
        return name_score + 2 * fill + pos

    best = max(numeric_cols, key=score)
    return best


def infer_text_column(df: pd.DataFrame) -> Optional[str]:
    text_cols = []
    for c in df.columns:
        if df[c].dtype == object:
            nonnull = df[c].astype(str).replace("nan", np.nan).notna().mean()
            if nonnull > 0.3:
                text_cols.append(c)

    if not text_cols:
        return None

    def score(col: str) -> float:
        name = str(col).lower()
        name_score = 0
        for kw in ["description", "particular", "item", "product", "service", "category", "narration", "vendor"]:
            if kw in name:
                name_score += 3
        return name_score

    return max(text_cols, key=score)


CAPEX_KEYWORDS = [
    "capex",
    "capital",
    "asset",
    "fixed asset",
    "equipment",
    "machinery",
    "machine",
    "computer",
    "laptop",
    "server",
    "printer",
    "furniture",
    "fixture",
    "vehicle",
    "construction",
    "renovation",
    "installation",
    "plant",
    "generator",
    "air conditioner",
    "ac ",
]


def detect_capital_goods(text: str) -> bool:
    t = (text or "").lower()
    return any(kw in t for kw in CAPEX_KEYWORDS)


def map_naics(
    df: pd.DataFrame,
    ref: ReferenceData,
    category_col: Optional[str],
    product_col: Optional[str],
    description_col: Optional[str],
) -> pd.DataFrame:
    out = df.copy()

    # If NAICS already present
    naics_col = None
    for c in out.columns:
        if "naics" in str(c).lower() and "code" in str(c).lower():
            naics_col = c
            break
    if naics_col is not None:
        out["naics_code"] = pd.to_numeric(out[naics_col], errors="coerce")
        out["naics_mapping_method"] = "from_input"
        return out

    # If Category+Product mapping is available from reference
    if (
        ref.catprod_to_naics is not None
        and category_col is not None
        and product_col is not None
        and category_col in out.columns
        and product_col in out.columns
    ):
        tmp = out.merge(
            ref.catprod_to_naics,
            how="left",
            left_on=[category_col, product_col],
            right_on=["category", "product"],
        )
        tmp["naics_mapping_method"] = np.where(tmp["naics_code"].notna(), "category_product_map", "unmapped")
        tmp = tmp.drop(columns=[c for c in ["category", "product"] if c in tmp.columns], errors="ignore")
        return tmp

    # Fuzzy match fallback against NAICS names (no external data; only your provided NAICS list)
    if rf_process is None:
        out["naics_code"] = np.nan
        out["naics_mapping_method"] = "unmapped"
        return out

    choices = (
        ref.naics_lookup.dropna(subset=["naics_code", "naics_name"]).assign(
            _choice=lambda d: d["naics_code"].astype(int).astype(str) + " | " + d["naics_name"].astype(str)
        )
    )
    choice_list = choices["_choice"].tolist()
    choice_to_code = dict(zip(choices["_choice"], choices["naics_code"]))

    def row_text(r) -> str:
        parts = []
        for col in [category_col, product_col, description_col]:
            if col and col in out.columns:
                v = r.get(col)
                if pd.notna(v):
                    parts.append(str(v))
        return " | ".join(parts)[:300]

    mapped_codes = []
    methods = []
    confidences = []

    for _, r in out.iterrows():
        txt = row_text(r)
        if not txt.strip():
            mapped_codes.append(np.nan)
            methods.append("unmapped")
            confidences.append(0)
            continue
        best = rf_process.extractOne(txt, choice_list, scorer=rf_fuzz.token_set_ratio)
        if not best:
            mapped_codes.append(np.nan)
            methods.append("unmapped")
            confidences.append(0)
            continue
        choice, score, _ = best
        mapped_codes.append(choice_to_code.get(choice))
        methods.append("fuzzy_naics_title")
        confidences.append(score)

    out["naics_code"] = mapped_codes
    out["naics_mapping_method"] = methods
    out["naics_mapping_confidence"] = confidences
    return out


def enrich_with_factors(df: pd.DataFrame, ref: ReferenceData) -> pd.DataFrame:
    out = df.copy()
    out = out.merge(ref.naics_lookup, how="left", on="naics_code")
    return out


def build_errors(df: pd.DataFrame) -> pd.Series:
    errs = []
    for _, r in df.iterrows():
        e = []
        if pd.isna(r.get("naics_code")):
            e.append("NAICS_NOT_MAPPED")
        if pd.isna(r.get("ef_kgco2e_per_usd_margin")):
            e.append("EF_NOT_FOUND")
        errs.append(";".join(e))
    return pd.Series(errs, index=df.index)


def to_excel_with_formulas(
    cat1: pd.DataFrame,
    cat2: pd.DataFrame,
    ref: ReferenceData,
    fx_inr_per_usd: float,
) -> bytes:
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # EF & Conversion sheet (copied as values)
    ws_ef = wb.create_sheet("EF & Conversion")

    # Write EF table header + first 8 columns (A:H) exactly; keep M1 for FX.
    ef_df = ref.ef_table.copy()

    # Ensure first 8 columns exist
    ef_export = ef_df.iloc[:, :8]
    # Header
    ws_ef.append(list(ef_export.columns))
    for row in ef_export.itertuples(index=False):
        ws_ef.append(list(row))

    # Put FX in M1 like the reference workbook does
    ws_ef["L1"].value = "INR per USD (manual)"
    ws_ef["M1"].value = float(fx_inr_per_usd)

    # Formatting header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_ef[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Output sheets
    def write_output_sheet(name: str, df: pd.DataFrame, highlight_mask: Optional[pd.Series] = None):
        ws = wb.create_sheet(name)
        headers = [
            "Goods or services",
            "NAICS code",
            "NAICS name",
            "Amount (INR)",
            "Converted amount (USD)",
            "Emission factor (kg CO2e/2022 USD) - with margin",
            "Emission (tCO2e)",
            "Errors",
        ]
        ws.append(headers)

        # Column widths (rough)
        widths = [40, 12, 55, 18, 22, 28, 16, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Header format
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        warn_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
        err_fill = PatternFill("solid", fgColor="F8CBAD")   # light red

        start_row = 2
        for i, r in enumerate(df.itertuples(index=False), start=start_row):
            goods_services = getattr(r, "goods_or_services")
            naics_code = getattr(r, "naics_code")
            amount_inr = getattr(r, "amount_inr")
            errors = getattr(r, "errors")

            ws.cell(i, 1).value = None if pd.isna(goods_services) else str(goods_services)
            ws.cell(i, 2).value = None if pd.isna(naics_code) else float(naics_code)

            # NAICS name via VLOOKUP to EF & Conversion (col 2)
            ws.cell(i, 3).value = f"=IFERROR(VLOOKUP(B{i},'EF & Conversion'!$A:$H,2,FALSE),\"\")"

            ws.cell(i, 4).value = None if pd.isna(amount_inr) else float(amount_inr)

            # USD conversion uses EF & Conversion!$M$1 (INR per USD)
            ws.cell(i, 5).value = f"=IFERROR(D{i}/'EF & Conversion'!$M$1,\"\")"

            # EF with margins is col 7 in EF & Conversion
            ws.cell(i, 6).value = f"=IFERROR(VLOOKUP(B{i},'EF & Conversion'!$A:$H,7,FALSE),\"\")"

            # Emissions in tCO2e
            ws.cell(i, 7).value = f"=IFERROR((E{i}*F{i})/1000,\"\")"

            ws.cell(i, 8).value = "" if pd.isna(errors) else str(errors)

            # Highlight rows with issues
            has_err = bool(ws.cell(i, 8).value)
            if has_err and "NOT_FOUND" in ws.cell(i, 8).value:
                for c in range(1, 9):
                    ws.cell(i, c).fill = err_fill
            elif has_err:
                for c in range(1, 9):
                    ws.cell(i, c).fill = warn_fill

            if highlight_mask is not None and bool(highlight_mask.iloc[i - start_row]):
                for c in range(1, 9):
                    ws.cell(i, c).fill = warn_fill

        return ws

    cat1_mask = cat1.get("flag_capital_goods_in_cat1", pd.Series([False] * len(cat1)))
    write_output_sheet("Scope3_Cat1", cat1, highlight_mask=cat1_mask)
    write_output_sheet("Scope3_Cat2", cat2, highlight_mask=None)

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def main():
    st.set_page_config(page_title="Scope 3 Spend-Based (NAICS) Tool", layout="wide")

    st.title("Scope 3 Spend-Based Emissions Tool (Category 1 & 2)")
    st.caption("Uses NAICS spend-based factors (with margins) from the provided reference workbook. Offline / no paid APIs.")

    # --- Reference workbook ---
    st.sidebar.header("Reference workbook")
    ref_upload = st.sidebar.file_uploader(
        "Upload reference Excel (required unless SCOPE3_REF_XLSX_PATH is set)",
        type=["xlsx"],
    )
    if ref_upload is not None:
        ref_path = ref_upload
        ref_label = "(uploaded)"
    else:
        ref_path = resolve_default_reference_path()
        ref_label = ref_path or "(no default found)"

    if ref_path is None:
        st.error(
            "Reference workbook is required in this hosted app. "
            "Upload it from the left sidebar under 'Reference workbook'."
        )
        st.stop()
        st.stop()
    try:
        ref = load_reference(ref_path)
    except Exception as e:
        st.error(f"Failed to load reference workbook {ref_label}: {e}")
        st.stop()

    # Decision 1: UI = Streamlit (already). Decision 2: No open-source model used by default.
    with st.sidebar.expander("Decisions (auto)", expanded=True):
        st.write("**UI framework:** Streamlit")
        st.write("**Open-source model:** Not used by default (mapping comes from your workbook + deterministic logic).")

    # --- Home inputs ---
    st.subheader("Home")
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        fy = st.text_input("Financial year (label)", value="FY 2025")
    with col2:
        default_fx = ref.default_fx_inr_per_usd or 85.0
        fx_inr_per_usd = st.number_input("INR per 1 USD", min_value=0.0001, value=float(default_fx), step=0.01)
    with col3:
        st.info(
            "The export will keep formulas (USD conversion and emissions) in Excel cells. "
            "FX is stored in 'EF & Conversion'!M1 like your reference workbook."
        )

    purchases_file = st.file_uploader("Upload purchases Excel", type=["xlsx"], key="purchases")
    if purchases_file is None:
        st.stop()

    file_kind = st.radio(
        "This upload is primarily for",
        ["Scope 3 Category 1 (Purchased goods & services)", "Scope 3 Category 2 (Capital goods)"],
        horizontal=True,
    )

    # List sheets
    try:
        wb = load_workbook(purchases_file, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
    except Exception as e:
        st.error(f"Could not open purchases workbook: {e}")
        st.stop()

    suggested_sheet = None
    for s in sheet_names:
        if str(s).strip().lower() in {"inward", "purchases", "purchase", "procurement"}:
            suggested_sheet = s
            break
    if suggested_sheet is None:
        suggested_sheet = sheet_names[0]

    sheet = st.selectbox("Select purchase sheet", options=sheet_names, index=sheet_names.index(suggested_sheet))

    # Load purchases sheet
    try:
        df = pd.read_excel(purchases_file, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read purchases sheet '{sheet}': {e}")
        st.stop()

    df = df.dropna(how="all")
    if df.empty:
        st.warning("Selected sheet is empty.")
        st.stop()

    # Infer key columns
    spend_col = infer_spend_column(df)
    desc_col = infer_text_column(df)

    # Category/Product columns if present
    category_col = next((c for c in df.columns if str(c).strip().lower() == "category"), None)
    product_col = next((c for c in df.columns if str(c).strip().lower() == "product"), None)

    with st.expander("Detected inputs (auto)", expanded=True):
        st.write({
            "rows": int(len(df)),
            "spend_column": str(spend_col),
            "description_column": str(desc_col) if desc_col else None,
            "category_column": str(category_col) if category_col else None,
            "product_column": str(product_col) if product_col else None,
        })

    # Build a working dataframe with required fields
    work = df.copy()
    work["amount_inr"] = pd.to_numeric(work[spend_col], errors="coerce")

    # Goods or services label (simple, editable later)
    if product_col is not None:
        work["goods_or_services"] = work[product_col].astype(str)
    elif desc_col is not None:
        work["goods_or_services"] = work[desc_col].astype(str)
    else:
        work["goods_or_services"] = "(unspecified)"

    # Capital goods detection
    text_for_capex = ""
    for c in [category_col, product_col, desc_col]:
        if c is not None and c in work.columns:
            text_for_capex += work[c].astype(str) + " | "
    work["is_capital_goods_detected"] = text_for_capex.apply(detect_capital_goods)

    # Map NAICS
    mapped = map_naics(work, ref, category_col=category_col, product_col=product_col, description_col=desc_col)
    mapped = enrich_with_factors(mapped, ref)

    # Optional mapping review / manual override
    enable_override = st.checkbox(
        "Review / override NAICS mapping (recommended when mappings are uncertain)",
        value=False,
    )
    if enable_override:
        st.write("Edit NAICS codes below. The tool will re-lookup NAICS name and EF from your reference workbook.")
        map_view = (
            mapped[["goods_or_services", "naics_code", "naics_name", "naics_mapping_method"]]
            .drop_duplicates(subset=["goods_or_services"])
            .sort_values("goods_or_services")
            .reset_index(drop=True)
        )
        edited = st.data_editor(
            map_view,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="naics_override",
        )
        override = edited[["goods_or_services", "naics_code"]].copy()
        override["naics_code"] = pd.to_numeric(override["naics_code"], errors="coerce")
        override = override.dropna(subset=["goods_or_services", "naics_code"])
        override_map = dict(zip(override["goods_or_services"].astype(str), override["naics_code"]))
        if override_map:
            mapped["naics_code"] = mapped.apply(
                lambda r: override_map.get(str(r.get("goods_or_services")), r.get("naics_code")), axis=1
            )
            base = mapped.drop(columns=["naics_name", "ef_kgco2e_per_usd_margin"], errors="ignore")
            mapped = enrich_with_factors(base, ref)
            mapped["naics_mapping_method"] = mapped["naics_mapping_method"].astype(str) + "|manual_override"

    # Build errors
    mapped["errors"] = build_errors(mapped)

    # Category split / flagging behavior
    if file_kind.startswith("Scope 3 Category 1"):
        # Keep everything in Cat1, but flag any rows that look like capital goods (to avoid miscounting).
        cat1 = mapped.copy()
        cat1["flag_capital_goods_in_cat1"] = cat1["is_capital_goods_detected"].fillna(False)
        cat2 = mapped[mapped["is_capital_goods_detected"]].copy()
    else:
        cat2 = mapped.copy()
        cat1 = mapped.iloc[0:0].copy()
        cat1["flag_capital_goods_in_cat1"] = pd.Series([], dtype=bool)
    # Display tabs
    tab_home, tab_cat1, tab_cat2, tab_diag = st.tabs(["Home", "Scope 3 Cat 1", "Scope 3 Cat 2 (Capital Goods)", "Diagnostics"])

    with tab_home:
        st.write("### Summary")
        st.write(
            f"**FY:** {fy}  |  **FX (INR/USD):** {fx_inr_per_usd:.4f}  |  "
            f"**Rows:** {len(mapped)}  |  **Cat1:** {len(cat1)}  |  **Cat2:** {len(cat2)}"
        )

        # Show a small preview
        st.write("### Preview (first 50 rows)")
        preview_cols = ["goods_or_services", "naics_code", "naics_name", "amount_inr", "ef_kgco2e_per_usd_margin", "errors", "naics_mapping_method"]
        st.dataframe(mapped[preview_cols].head(50))

        # Export
        out_bytes = to_excel_with_formulas(
            cat1[["goods_or_services", "naics_code", "amount_inr", "errors", "flag_capital_goods_in_cat1"]].copy(),
            cat2[["goods_or_services", "naics_code", "amount_inr", "errors"]].copy(),
            ref,
            fx_inr_per_usd=float(fx_inr_per_usd),
        )

        st.download_button(
            label="Download output Excel (with formulas)",
            data=out_bytes,
            file_name="scope3_spend_based_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with tab_cat1:
        st.write("### Scope 3 Category 1 output")
        st.dataframe(
            cat1[["goods_or_services", "naics_code", "naics_name", "amount_inr", "ef_kgco2e_per_usd_margin", "errors", "naics_mapping_method"]]
        )

    with tab_cat2:
        st.write("### Scope 3 Category 2 (Capital goods) output")
        st.dataframe(
            cat2[["goods_or_services", "naics_code", "naics_name", "amount_inr", "ef_kgco2e_per_usd_margin", "errors", "naics_mapping_method"]]
        )

    with tab_diag:
        st.write("### Unmapped / missing factor rows")
        diag = mapped[(mapped["errors"] != "")].copy()
        st.dataframe(diag[["goods_or_services", "naics_code", "naics_name", "amount_inr", "errors", "naics_mapping_method"]])


if __name__ == "__main__":
    main()
